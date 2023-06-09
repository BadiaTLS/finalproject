from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import user_passes_test
from django.contrib import messages
from .utils import *
import openpyxl
from final_project.accounts.models import CustomUser
from .models import table_classes, Attendee

# Create your views here.
def check_sas_role(user):
    return user.role == 'sas'

@login_required(login_url='login')
@user_passes_test(check_sas_role, login_url='not_sas')
def sas_index(request):
    user_data = get_all_user()
    class_data = get_all_class()
    return render(request, "sas/sas_index.html", {'user_data':user_data, 'class_data':class_data})

@login_required(login_url='login')
@user_passes_test(check_sas_role, login_url='not_sas')
def import_user(request):
    if request.method == 'POST':
        excel_file = request.FILES['excel_file']

        if excel_file.name.endswith('.xlsx'):
            workbook = openpyxl.load_workbook(excel_file)
            worksheet = workbook.active

            for row in worksheet.iter_rows(min_row=2, values_only=True):
                user_id = row[0]
                name = row[1]
                email = row[2]
                role = row[3]
                
                # Check if user with the same email exists
                user = CustomUser.objects.filter(email=email).first()
                if user:
                    # User already exists, update the fields
                    user.username = user_id
                    user.name = name
                    user.role = role
                    user.save()
                else:
                    # Create a new user
                    user = CustomUser.objects.create_user(username=user_id, password=user_id, email=email)
                    user.name = name
                    user.role = role
                    user.save()

            messages.success(request, 'Users imported successfully.', extra_tags='success')
            return redirect('sas_index')
        else:
            messages.error(request, 'Invalid file format. Please upload an Excel file.', extra_tags='error')

    return render(request, 'sas/import_user.html')

@login_required(login_url='login')
@user_passes_test(check_sas_role, login_url='not_sas')
def import_class(request):
    if request.method == 'POST':
        excel_file = request.FILES['excel_file']

        if excel_file.name.endswith('.xlsx'):
            workbook = openpyxl.load_workbook(excel_file)
            worksheet = workbook.active

            for row in worksheet.iter_rows(min_row=5, values_only=True):
                class_code = row[0]
                class_name = row[1]
                class_day = row[2]
                class_start_time = row[3]
                class_end_time = row[4]
                list_attendees = row[5].split(",")  # Split attendees by comma

                # Retrieve or create the class instance
                table_class, _ = table_classes.objects.get_or_create(
                    class_code=class_code,
                    defaults={
                        'class_name': class_name,
                        'class_day': class_day,
                        'class_start_time': class_start_time,
                        'class_end_time': class_end_time,
                    }
                )

                # Add attendees to the class instance
                for attendee_email in list_attendees:
                    attendee, _ = Attendee.objects.get_or_create(email=attendee_email.strip())
                    table_class.attendees.add(attendee)

            messages.success(request, 'Classes imported successfully.', extra_tags='success')
            return redirect('sas_index')
        else:
            messages.error(request, 'Invalid file format. Please upload an Excel file (.xlsx).', extra_tags='error')
            return redirect('import_class')
    else:
        return render(request, 'sas/import_class.html')

def not_sas(request):
    messages.error(request, 'You are not authorized to access different role resources', extra_tags='error')
    return redirect('sas_index')
